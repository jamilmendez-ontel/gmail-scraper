#!/usr/bin/env python3
"""
Email notification module — sends pipeline reports with Excel attachment.

Sends via the pipeline's nanoninth Gmail account (separate credentials from
the scraper's Ontel account used for reading).
"""

import base64
import pickle
import re
import traceback
from io import BytesIO
from pathlib import Path
from datetime import datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from typing import List

from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from openpyxl import Workbook
from openpyxl.styles import numbers

from config import SCHEMA_STAGING, REPORT_EMAIL_TO, NOTIFY_CREDENTIALS_DIR, get_logger

SCHEMA_ANALYTICS = "analytics"
from db import get_db

logger = get_logger("notifier")

# Date formats found in package email fields (tried in order — most specific first)
_DATE_FORMATS = [
    "%m-%d-%Y %I:%M %p",  # 02-25-2026 01:40 PM
    "%m/%d/%Y %I:%M %p",  # 10/16/2025 2:55 PM
    "%m/%d/%y %I:%M %p",  # 2/26/26 3:00 PM
    "%m-%d-%Y",            # 02-26-2026
    "%m/%d/%Y",            # 2/3/2026, 12/22/2025
    "%m/%d/%y",            # 2/27/26
]

# Columns that contain date values
_DATE_COLUMNS = {
    "raw_files_received", "cx_start", "cx_complete",
    "live_review_complete", "revision_files_received",
    "revision_complete", "cop_complete",
    "cutover_complete", "hr48_raw_files_received", "hr48_package_complete",
    "pmi_cop_complete", "ll_cop_complete",
}


def _try_parse_date(val: str):
    """Try to parse a string as a date. Returns datetime on success, empty string for placeholders, original string otherwise."""
    if not val:
        return val
    cleaned = val.strip()

    # Skip known non-date placeholders
    if cleaned in ("N/A", "No", "PENDING ITEMS", "- -", "--/--/----", "--", ""):
        return ""

    # Skip time-only values (no date component, e.g. "12:00:00 AM")
    if re.match(r'^\d{1,2}:\d{2}(:\d{2})?\s*(AM|PM)?$', cleaned, re.IGNORECASE):
        return ""

    # Remove stray spaces from date values (e.g. "02-1 9 -2026" → "02-19-2026")
    # Split off AM/PM time portion first to preserve its spacing
    m = re.match(r'^(.*?)(\d{1,2}:\d{2}(?::\d{2})?\s*(?:AM|PM))$', cleaned, re.IGNORECASE)
    if m:
        date_part = m.group(1).replace(' ', '')
        time_part = m.group(2).strip()
        cleaned = f"{date_part} {time_part}"
    else:
        cleaned = cleaned.replace(' ', '')

    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(cleaned, fmt)
        except ValueError:
            continue
    return val

# Scopes needed by the sender account (must match the pipeline's token)
_SEND_SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/gmail.send",
]


def _get_sender_service():
    """
    Authenticate using the pipeline's nanoninth Gmail credentials.

    Returns an authorized Gmail API service for sending.
    """
    creds_dir = Path(NOTIFY_CREDENTIALS_DIR)
    token_file = creds_dir / "token.pickle"

    if not token_file.exists():
        raise FileNotFoundError(
            f"Sender token not found at {token_file}. "
            "Ensure the local-pipeline has been authenticated first."
        )

    with open(token_file, "rb") as f:
        creds = pickle.load(f)

    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
        with open(token_file, "wb") as f:
            pickle.dump(creds, f)

    return build("gmail", "v1", credentials=creds)


def generate_excel(thread_ids: List[str]) -> bytes:
    """
    Query stg_package_emails + stg_emails for given thread_ids and
    return an in-memory Excel workbook as bytes.

    Dynamic columns: fixed columns from the view come first, then any
    extra JSONB field keys discovered at query time are appended
    automatically — no code change needed when new fields appear.
    """
    if not thread_ids:
        return _empty_workbook()

    db = get_db()

    # Query the deduped view for fixed columns + raw fields JSONB
    rows = db.fetch(
        f"""
        SELECT v.*, c.fields
        FROM {SCHEMA_ANALYTICS}.v_package_emails v
        JOIN {SCHEMA_STAGING}.stg_package_emails c USING (message_id)
        WHERE v.thread_id = ANY($1::text[])
        ORDER BY v.received_at_et
        """,
        thread_ids,
    )

    # Fixed columns (always present, in this order)
    fixed_columns = [
        ("received_at_et", "Received (ET)"),
        ("sender_email", "Sender"),
        ("clean_subject", "Subject"),
        ("subject", "Raw Subject"),
        ("package_type", "Package Type"),
        ("site_id", "Site ID"),
        ("site_name", "Site Name"),
        ("gc_name", "GC Name"),
        ("landlord", "Landlord"),
        ("project", "Project"),
        ("project_id", "Project ID"),
        ("market", "Market"),
        ("structure_type", "Structure Type"),
        ("cm_company", "CM Company"),
        ("cm_name", "CM Name"),
        ("project_manager", "Project Manager"),
        ("equipment_engineer", "Equipment Engineer"),
        ("construction_engineer", "Construction Engineer"),
        ("raw_files_received", "Raw Files Received"),
        ("cx_start", "CX Start"),
        ("cx_complete", "CX Complete"),
        ("cx_duration", "CX Duration"),
        ("live_review_complete", "Live Review Complete"),
        ("live_review_duration", "Live Review Duration"),
        ("revision_files_received", "Revision Files Received"),
        ("revision_complete", "Revision Complete"),
        ("cop_complete", "COP Complete"),
        ("cop_status", "COP Status"),
        ("cop_duration", "COP Duration"),
        ("cop_raw_file_duration", "COP Raw File Duration"),
        ("cutover_complete", "Cutover Complete"),
        ("hr48_raw_file_duration", "48Hr Raw File Duration"),
        ("hr48_package_duration", "48Hr Package Duration"),
        ("hr48_raw_files_received", "48Hr Raw Files Received"),
        ("hr48_package_complete", "48Hr Package Complete"),
        ("pmi_cop_complete", "PMI COP Complete"),
        ("smart_tool_project_num", "Smart Tool Project #"),
        ("mdg_location_id", "MDG Location ID"),
        ("landlord_site_name", "Landlord Site Name"),
        ("ll_cop_complete", "LL COP Complete"),
        ("open_items", "Open Items"),
        ("dropbox_url", "Dropbox URL"),
        ("swift_url", "Swift URL"),
    ]

    # Keys already covered by the fixed columns (normalized from JSONB)
    _known_field_keys = {
        "SITE ID", "Site ID", "Landlord Site ID",
        "SITE NAME", "Site Name", "Carrier Site Name",
        "GC NAME", "GC Name",
        "LANDLORD", "Landlord",
        "PROJECT", "Project Type", "Carrier Project Type",
        "PROJECT ID", "Project ID", "Carrier Project ID",
        "MARKET", "Market-County", "County",
        "STRUCTURE TYPE", "Structure Type",
        "CM Company", "CM Name",
        "Project Manager", "Equipment Engineer",
        "Construction Engineer", "A&E Company",
        "RAW FILES RECEIVED", "COP Raw Files Received",
        "CX START", "CX Start",
        "CX COMPLETE", "CX Complete",
        "CX Duration",
        "LIVE REVIEW COMPLETE", "Live Review Complete",
        "Live Review Duration",
        "REVISION FILES RECEIVED", "REVISION COMPLETE",
        "Revision Files Received", "Revision Complete",
        "COP COMPLETE", "COP Complete",
        "COP Status", "COP Duration", "COP Raw File Duration",
        "Cutover Complete",
        "48Hr Raw File Duration", "48Hr Package Duration",
        "48Hr Raw Files Received", "48Hr Package Complete",
        "PMI COP Complete",
        "Smart Tool Project #", "MDG Location ID",
        "Landlord Site Name",
        "LL COP Complete", "LL COP COMPLETE",
        "Open Items",
    }

    # Discover any new JSONB keys not covered by fixed columns
    extra_keys: list[str] = []
    seen_extra: set[str] = set()
    for row in rows:
        fields = row["fields"] if row["fields"] else {}
        for key in fields:
            if key not in _known_field_keys and key not in seen_extra:
                extra_keys.append(key)
                seen_extra.add(key)

    wb = Workbook()
    ws = wb.active
    ws.title = "Package Records"

    # Header row: fixed columns + any dynamic extras
    headers = [label for _, label in fixed_columns] + extra_keys
    ws.append(headers)

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        fields = row["fields"] if row["fields"] else {}
        col_idx = 1

        # Fixed columns
        for col_key, _ in fixed_columns:
            val = row[col_key]
            cell = ws.cell(row=row_idx, column=col_idx)

            if val is None:
                cell.value = ""
            elif hasattr(val, "strftime"):
                cell.value = val
                cell.number_format = "YYYY-MM-DD HH:MM"
            elif col_key in _DATE_COLUMNS:
                parsed = _try_parse_date(str(val))
                if isinstance(parsed, datetime):
                    cell.value = parsed
                    if parsed.hour or parsed.minute:
                        cell.number_format = "MM/DD/YYYY HH:MM AM/PM"
                    else:
                        cell.number_format = "MM/DD/YYYY"
                else:
                    cell.value = str(parsed)
            else:
                cell.value = str(val)
            col_idx += 1

        # Dynamic extra columns from JSONB
        for key in extra_keys:
            val = fields.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx)
            if val:
                parsed = _try_parse_date(str(val))
                if isinstance(parsed, datetime):
                    cell.value = parsed
                    if parsed.hour or parsed.minute:
                        cell.number_format = "MM/DD/YYYY HH:MM AM/PM"
                    else:
                        cell.number_format = "MM/DD/YYYY"
                else:
                    cell.value = str(val)
            else:
                cell.value = ""
            col_idx += 1

    # Auto-size columns (approximate)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _empty_workbook() -> bytes:
    """Return a minimal Excel file with just a header row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Package Records"
    ws.append(["No new package records this run"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_html_email(message_ids: List[str], started: datetime, ended: datetime) -> str:
    """Build an HTML email body matching the main pipeline's style."""
    from datetime import timedelta

    db = get_db()

    total_cop = db.fetchrow(
        f"SELECT COUNT(*) AS cnt FROM {SCHEMA_ANALYTICS}.v_package_emails"
    )
    total_scraped = db.fetchrow(
        f"SELECT COUNT(*) AS cnt FROM {SCHEMA_STAGING}.stg_package_emails"
    )
    total_emails = db.fetchrow(
        f"SELECT COUNT(*) AS cnt FROM {SCHEMA_STAGING}.stg_emails"
    )
    by_type = db.fetch(
        f"""
        SELECT package_type, COUNT(*) AS cnt
        FROM {SCHEMA_ANALYTICS}.v_package_emails
        GROUP BY package_type ORDER BY cnt DESC
        """
    )

    new_count = len(message_ids)
    status = "SUCCESS"
    banner_color = "#2e7d32"

    # Duration
    duration_s = int((ended - started).total_seconds())
    if duration_s >= 60:
        dur_str = f"{duration_s // 60}m {duration_s % 60}s"
    else:
        dur_str = f"{duration_s}s"

    # ET timestamps
    et_offset = timedelta(hours=-5)
    et_tz = timezone(et_offset)
    started_et = started.astimezone(et_tz).strftime("%Y-%m-%d %H:%M:%S EST")
    ended_et = ended.astimezone(et_tz).strftime("%Y-%m-%d %H:%M:%S EST")

    # Build package type rows
    type_rows = ""
    for row in by_type:
        type_rows += (
            f'<tr>'
            f'<td style="padding:6px 12px;border:1px solid #ddd;">{row["package_type"]}</td>'
            f'<td style="padding:6px 12px;border:1px solid #ddd;text-align:right;">{row["cnt"]}</td>'
            f'</tr>'
        )

    # Change color for new records
    if new_count > 0:
        change_color = "#2e7d32"
        change_str = f"+{new_count}"
    else:
        change_color = "#888"
        change_str = "0"

    html = f"""
    <html><body style="font-family:Arial,sans-serif;margin:0;padding:0;">
    <div style="background-color:{banner_color};color:white;padding:16px 24px;border-radius:4px 4px 0 0;">
        <h2 style="margin:0;">Gmail Package Scraper: {status}</h2>
    </div>
    <div style="padding:16px 24px;">

        <table style="margin-bottom:16px;">
            <tr><td style="padding:2px 16px 2px 0;font-weight:bold;">Started:</td><td>{started_et}</td></tr>
            <tr><td style="padding:2px 16px 2px 0;font-weight:bold;">Ended:</td><td>{ended_et}</td></tr>
            <tr><td style="padding:2px 16px 2px 0;font-weight:bold;">Duration:</td><td>{dur_str}</td></tr>
        </table>

        <h3 style="margin-top:24px;margin-bottom:8px;">Pipeline Details</h3>
        <table style="border-collapse:collapse;width:100%;">
            <thead>
                <tr style="background-color:#f5f5f5;">
                    <th style="padding:6px 12px;border:1px solid #ddd;text-align:left;">Step</th>
                    <th style="padding:6px 12px;border:1px solid #ddd;text-align:left;">Status</th>
                    <th style="padding:6px 12px;border:1px solid #ddd;text-align:left;">Details</th>
                </tr>
            </thead>
            <tbody>
                <tr>
                    <td style="padding:6px 12px;border:1px solid #ddd;">Gmail Scrape</td>
                    <td style="padding:6px 12px;border:1px solid #ddd;color:#2e7d32;font-weight:bold;">SUCCESS</td>
                    <td style="padding:6px 12px;border:1px solid #ddd;">{new_count} new emails fetched</td>
                </tr>
                <tr>
                    <td style="padding:6px 12px;border:1px solid #ddd;">Package Parser</td>
                    <td style="padding:6px 12px;border:1px solid #ddd;color:#2e7d32;font-weight:bold;">SUCCESS</td>
                    <td style="padding:6px 12px;border:1px solid #ddd;">{new_count} emails parsed</td>
                </tr>
            </tbody>
        </table>

        <h3 style="margin-top:24px;margin-bottom:8px;">Row Counts</h3>
        <table style="border-collapse:collapse;">
            <thead>
                <tr style="background-color:#f5f5f5;">
                    <th style="padding:6px 12px;border:1px solid #ddd;text-align:left;">Table</th>
                    <th style="padding:6px 12px;border:1px solid #ddd;text-align:right;">Total</th>
                    <th style="padding:6px 12px;border:1px solid #ddd;text-align:right;">New</th>
                </tr>
            </thead>
            <tbody>
                <tr>
                    <td style="padding:6px 12px;border:1px solid #ddd;">stg_emails</td>
                    <td style="padding:6px 12px;border:1px solid #ddd;text-align:right;">{total_emails['cnt']:,}</td>
                    <td style="padding:6px 12px;border:1px solid #ddd;text-align:right;color:{change_color};font-weight:bold;">{change_str}</td>
                </tr>
                <tr>
                    <td style="padding:6px 12px;border:1px solid #ddd;">stg_package_emails</td>
                    <td style="padding:6px 12px;border:1px solid #ddd;text-align:right;">{total_scraped['cnt']:,}</td>
                    <td style="padding:6px 12px;border:1px solid #ddd;text-align:right;color:{change_color};font-weight:bold;">{change_str}</td>
                </tr>
                <tr>
                    <td style="padding:6px 12px;border:1px solid #ddd;">v_package_emails (deduped)</td>
                    <td style="padding:6px 12px;border:1px solid #ddd;text-align:right;">{total_cop['cnt']:,}</td>
                    <td style="padding:6px 12px;border:1px solid #ddd;text-align:right;">-</td>
                </tr>
            </tbody>
        </table>

        <h3 style="margin-top:24px;margin-bottom:8px;">Records by Package Type</h3>
        <table style="border-collapse:collapse;">
            <thead>
                <tr style="background-color:#f5f5f5;">
                    <th style="padding:6px 12px;border:1px solid #ddd;text-align:left;">Package Type</th>
                    <th style="padding:6px 12px;border:1px solid #ddd;text-align:right;">Count</th>
                </tr>
            </thead>
            <tbody>
                {type_rows}
            </tbody>
        </table>

    </div>
    </body></html>
    """
    return html


def send_report(log_text: str, message_ids: List[str],
                started: datetime = None, ended: datetime = None):
    """
    Build and send the pipeline report email via the nanoninth account.

    Attachments:
      - package_records_YYYY-MM-DD.xlsx (parsed package data from analytics view)
      - scraper_YYYY-MM-DD.log (full pipeline log output)

    Wrapped in try/except — email failures are logged but never crash the pipeline.
    """
    if not REPORT_EMAIL_TO:
        logger.info("REPORT_EMAIL_TO not set — skipping notification email")
        return

    if ended is None:
        ended = datetime.now(timezone.utc)
    if started is None:
        started = ended

    try:
        service = _get_sender_service()

        # Resolve message_ids to thread_ids for deduped view filtering
        db = get_db()
        thread_ids = []
        if message_ids:
            rows = db.fetch(
                f"""
                SELECT DISTINCT thread_id
                FROM {SCHEMA_STAGING}.stg_emails
                WHERE message_id = ANY($1::text[])
                """,
                message_ids,
            )
            thread_ids = [r["thread_id"] for r in rows]

        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        subject = f"Gmail Package Scraper: SUCCESS -- {today}"

        # Build MIME message
        msg = MIMEMultipart()
        msg["To"] = REPORT_EMAIL_TO
        msg["Subject"] = subject

        # HTML body
        html_body = _build_html_email(message_ids, started, ended)
        msg.attach(MIMEText(html_body, "html"))

        # Excel attachment (only records from this run, filtered by thread)
        excel_bytes = generate_excel(thread_ids)
        excel_filename = f"package_records_{today}.xlsx"
        excel_part = MIMEBase(
            "application",
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        excel_part.set_payload(excel_bytes)
        encoders.encode_base64(excel_part)
        excel_part.add_header("Content-Disposition", "attachment", filename=excel_filename)
        msg.attach(excel_part)

        # Log file attachment
        log_filename = f"scraper_{today}.log"
        log_part = MIMEBase("text", "plain")
        log_part.set_payload(log_text.encode("utf-8"))
        encoders.encode_base64(log_part)
        log_part.add_header("Content-Disposition", "attachment", filename=log_filename)
        msg.attach(log_part)

        # Send
        raw = base64.urlsafe_b64encode(msg.as_string().encode()).decode()
        service.users().messages().send(
            userId="me", body={"raw": raw}
        ).execute()

        logger.info(f"Report email sent to {REPORT_EMAIL_TO}")

    except Exception as e:
        logger.error(f"Failed to send notification email: {e}\n{traceback.format_exc()}")
